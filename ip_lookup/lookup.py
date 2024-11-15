"""
    Implement IP lookup, using RIR or local cache
"""
import time
import socket
import ipaddress
import logging
import re

from enum import Enum

import pandas
import ipwhois
import openpyxl
import openpyxl.styles as xlstyle

from ip_lookup.cache import NetworkCache, ResolvedNetwork, DEFAULT_CACHE_FILE

RETRY_COUNT_MAX: int = 3
SLEEP_TIME = 2
SLEEP_INTERVAL = 10
CACHE_SAVE_INTERVAL = 100
TIMEOUT_SLEEP = 10
EXCEL_ENTRIES_MARKER = "Entries"
EXCEL_ENTRIES_MAX_SEARCH = 10
EXCEL_MAX_COLS = 100

ReservedNetNames = [
    "IANA-BLK"
]


class Descriptions(Enum):
    """
        Descriptions for well-known networks, such as private, loopback, etc.
    """
    PRIVATE_IP_DESCR = "Private (RFC 1918 or APIPA) range"
    LOOPBACK_IP_DESCR = "Loopback range"
    MCAST_IP_DESCR = "Multicast range"
    RSVD_IP_DESCR = "Reserved IP range"
    INVALID_IP_DESCR = "Invalid IP Address"


class RirSearcher:
    """
        Performs RIR or cache lookup for addresses from text file or
        SSL inspection spreadsheet
    """
    resolved_ip_list: list[ResolvedNetwork] = []
    raw_ip_list: list[str] = []
    cache: NetworkCache = None
    start_cache_size: int = 0

    def __init__(self, ip_file_name: str = None, cache_file_name: str = DEFAULT_CACHE_FILE):
        self.cache = NetworkCache(cache_file_name)
        self.start_cache_size = len(self.cache.cache)

        if ip_file_name is not None:
            with open(ip_file_name, "r", encoding="utf-8") as file:
                # Remove /n part of the address along with whitespaces
                self.raw_ip_list: list[str] = [line.replace("\n", "").replace(" ", "")
                                               for line in file]

    def reload_file(self, ip_file_name: str) -> None:
        """
            Load file with the list of IP addresses, replacing the previous list

            :param str ip_file_name: path to the file with IP list
        """
        with open(ip_file_name, "r", encoding="utf-8") as file:
            # Remove /n part of the address along with whitespaces
            self.raw_ip_list: list[str] = [line.replace("\n", "").replace(" ", "")
                                           for line in file]

    def _check_known_nets(self, address: str)\
            -> tuple[ipaddress.IPv4Address | None, ResolvedNetwork | None]:
        """
            Check if the requested IP address or network match well-known values.

            :param str address: IP address or network in CIDR notation

            :return IPv4Address: object for IP or host part of subnet; None if well-known
            :return ResolvedNetwork: None, if lookup is required; resolved object if well-known
        """
        host_addr: str = address.split("/")[0]
        try:
            ip: ipaddress.IPv4Address = ipaddress.ip_address(host_addr)
        except ValueError:
            return None, ResolvedNetwork(address=address,
                                         description=Descriptions.INVALID_IP_DESCR.value)

        # Exclude well-known addresses
        if ip.is_private:
            return None, ResolvedNetwork(address=address,
                                         description=Descriptions.PRIVATE_IP_DESCR.value)
        if ip.is_loopback:
            return None, ResolvedNetwork(address=address,
                                         description=Descriptions.LOOPBACK_IP_DESCR.value)
        if ip.is_multicast:
            return None, ResolvedNetwork(address=address,
                                         description=Descriptions.MCAST_IP_DESCR.value)
        if ip.is_unspecified or ip.is_reserved:
            return None, ResolvedNetwork(address=address,
                                         description=Descriptions.RSVD_IP_DESCR.value)

        return ip, None

    def single_lookup(self, address: str, persist: bool = True)\
            -> tuple[bool, bool, bool, ResolvedNetwork | None]:
        """
            Perform an address resolution for a single entry.

            :param str address: IP address or network in CIDR notation
            :param bool persist: True, if cache needs to be dumped to disk

            :return bool known_net: is network a well-known
            :return bool cached: is found in cache
            :return bool error: has error occurred during lookup
            :return ResolvedNetwork resolved_net: None if error, resolved entry otherwise
        """
        ip, resolved_net = self._check_known_nets(address)

        if resolved_net is not None:
            return True, False, False, resolved_net

        # Check if IP in cache
        if net := self.cache.in_cache(ip):
            data = self.cache.get_network(address, net)
            return False, True, False, data

        whois: ipwhois.IPWhois = ipwhois.IPWhois(str(ip))
        retry_count = 0
        response: dict = {}
        while retry_count < RETRY_COUNT_MAX:
            try:
                response = whois.lookup_whois()
                break
            except TimeoutError:
                # wait a few moments before trying again
                logging.info("Pausing lookup for %s seconds because of likely timeout error",
                             TIMEOUT_SLEEP)
                retry_count += 1
                time.sleep(TIMEOUT_SLEEP)
            except (ipwhois.exceptions.HTTPLookupError, ipwhois.exceptions.WhoisLookupError,
                    ipwhois.exceptions.ASNParseError) as e:
                logging.warning("Lookup failed for %s with exception %s", address, str(e))
                # response = {}
                break

        if retry_count == RETRY_COUNT_MAX or not response:
            self.cache.not_found.append(address)
            return False, False, True, None

        # Build description and append to output list
        if len(response["nets"]) > 0:
            net_cidr: str = str(response["nets"][0]["cidr"]).replace("\n", " ")
            net_name: str = str(response["nets"][0]["name"]).replace("\n", " ")
            net_description: str = str(response["nets"][0]["description"]).replace("\n", " ")
            net_country: str = str(response["nets"][0]["country"]).replace("\n", " ")
        else:
            net_cidr: str = str(response["asn_cidr"]).replace("\n", " ")
            net_name: str = ""
            net_description: str = str(response["asn_description"]).replace("\n", " ")
            net_country: str = str(response["asn_country_code"]).replace("\n", " ")

        net_registry: str = str(response["asn_registry"]).replace("\n", " ")
        net_rdns_fqdn: str = ""

        host_and_mask = address.split("/")
        if len(host_and_mask) == 1 or int(host_and_mask[1]) == 32:
            try:
                net_rdns_fqdn = socket.gethostbyaddr(host_and_mask[0])[0]
            except socket.herror:
                pass

        resolved_net = ResolvedNetwork(address=address, name=net_name, cidr=net_cidr,
                                      country=net_country, registry=net_registry,
                                      description=net_description, fqdn=net_rdns_fqdn)

        # If network is too wide and delegated to other RIR –
        # skip adding to cache and searching subnetworks
        if net_name not in ReservedNetNames:
            # Get subnets for the aggregate – either 'asn_cidr' or 'nets':first_net:'cidr'
            subnets = [response.get('asn_cidr', None)]
            if subnets[0] in [None, "NA"]:
                subnets = response['nets'][0]['cidr'].replace(" ", "").split(",")

            # Add to cache for subnet in response
            for net in subnets:
                self.cache.set(net, net_name, net_description,
                               net_country, net_registry, net_rdns_fqdn)

            if persist:
                self.cache.save_cache(DEFAULT_CACHE_FILE)
                self.cache.save_not_found()

        return False, False, False, resolved_net

    def search_list(self):
        """
            Search the list of IP addresses from the file. Filename can either be
            specified during class init or the list can be manually updated.
        """
        ip_list = self.raw_ip_list.copy()
        cache_hits = 0
        lookups = 0

        while len(ip_list) != 0:
            address: str = ip_list.pop(0)
            known_net, cached, error, resolved_net = self.single_lookup(address, persist=False)

            if known_net:
                self.resolved_ip_list.append(resolved_net)
                continue

            if cached:
                cache_hits += 1
                self.resolved_ip_list.append(resolved_net)
                continue

            # Perform lookup with RIR
            lookups += 1

            if error:
                self.cache.not_found.append(address)
                continue

            if lookups % SLEEP_INTERVAL == 0:
                logging.info("Sleeping %s seconds after another %s lookups. Total lookups: %s",
                             SLEEP_TIME, SLEEP_INTERVAL, lookups)
                time.sleep(SLEEP_TIME)

            if lookups % CACHE_SAVE_INTERVAL == 0:
                # save cache file to not lose progress
                self.cache.save_cache(DEFAULT_CACHE_FILE)

            self.resolved_ip_list.append(resolved_net)

        # Save cache if modified
        if len(self.cache.cache) > self.start_cache_size:
            self.cache.save_cache(DEFAULT_CACHE_FILE)

        self.cache.save_not_found()

        logging.info("%s Cache hits | %s Lookups", cache_hits, lookups)

    def search_excel(self, excel: str, sheet_list: list[str]) -> None:
        """
            Search CSE SSL spreadsheet for IPs and resolve them

            :param str excel: path to CSE SSL spreadsheet
            :param list[str] sheet_list: worksheet names to search through
        """
        cache_hits: int = 0
        lookups: int = 0
        align_multiline: xlstyle.Alignment = xlstyle.Alignment(wrapText=True)
        background_fill: xlstyle.PatternFill = (
            xlstyle.PatternFill(patternType='solid', fgColor=xlstyle.colors.Color(rgb='DDEBF7')))

        wb: openpyxl.Workbook = openpyxl.load_workbook(filename=excel)

        logging.info("Building lookup set from spreadsheets")
        for sheet in sheet_list:
            ws = wb[sheet]
            entries_start_row: int = -1

            logging.info("Processing sheet %s", sheet)

            # Find the row that is used as the row for first entries in lists
            for row in range(1, EXCEL_ENTRIES_MAX_SEARCH):
                if ws.cell(row=row, column=1).value == EXCEL_ENTRIES_MARKER:
                    entries_start_row = row
                    break

            # if no starting row is found, go to next sheet
            if entries_start_row == -1:
                continue

            for col in range(2, EXCEL_MAX_COLS):
                # If column is empty, not further lists are available in the sheet
                if ws.cell(row=1, column=col).value == "":
                    break

                # Iterate over cells and resolve IPs
                row = entries_start_row - 1
                while True:
                    row += 1
                    entry = ws.cell(row=row, column=col).value
                    ws.cell(row=row, column=col).alignment = align_multiline

                    # If cell is empty or not defined, list has ended
                    if entry == "" or entry is None:
                        break

                    # Roughly select entries that resemble IPs or subnets
                    if re.match(r"^[\d\.\/]+$", entry) is None:
                        continue

                    address = entry.replace("\n", "")#.split("/")[0]

                    # If value cannot be converted to IPv4Address – skip
                    try:
                        known_net, cached, error, resolved_net =(
                            self.single_lookup(address=address, persist=False))
                    except ValueError:
                        continue

                    # Mark cell with colour as containing IP address
                    ws.cell(row=row, column=col).fill = background_fill

                    if error:
                        ws.cell(row=row, column=col).value = f"{entry}\n\nNOT FOUND"
                        continue

                    ws.cell(row=row, column=col).value = str(resolved_net)
                    lookups += 1

                    if known_net:
                        continue

                    if cached:
                        cache_hits += 1
                        continue

                    if lookups % SLEEP_INTERVAL == 0:
                        logging.info("Sleeping %s seconds after another %s lookups. "
                                     "Total lookups: %s",
                                     SLEEP_TIME, SLEEP_INTERVAL, lookups)
                        time.sleep(SLEEP_TIME)

                    if lookups % CACHE_SAVE_INTERVAL == 0:
                        # save cache file to not lose progress
                        self.cache.save_cache(DEFAULT_CACHE_FILE)

        logging.info("cache_hits=%s | lookups=%s", cache_hits, lookups)

        # Save cache if modified
        if len(self.cache.cache) > self.start_cache_size:
            self.cache.save_cache(DEFAULT_CACHE_FILE)

        self.cache.save_not_found()

        wb.save(excel)

    def to_excel(self, filename: str = "out.xlsx") -> None:
        """
            Export resolved IPs to Excel spreadsheet

            :param str filename: path for Excel spreadsheet
        """
        output: list[dict[str, str]] = []
        for x in self.resolved_ip_list:
            output.append(x.to_dict())

        with pandas.ExcelWriter(filename, engine="xlsxwriter") as xlwriter:
            df = pandas.DataFrame(output)
            df.to_excel(xlwriter, sheet_name="Data", index=False, header=True)
